from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi import UploadFile, File
import shutil
import uuid
from openpyxl import Workbook, load_workbook
from io import BytesIO
import sqlite3, json, os, requests, base64, httpx, asyncio
from starlette.responses import Response
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

from datetime import datetime

@app.middleware("http")
async def add_timestamp(request: Request, call_next):
    request.state.timestamp = int(datetime.now().timestamp())
    response = await call_next(request)
    return response

app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

DB_PATH = "quiz.db"

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# -------------------- ROUTES --------------------
@app.get("/", response_class=HTMLResponse)
@app.get("/home", response_class=HTMLResponse)
def home(request: Request):
    print(request)
    return templates.TemplateResponse(
        "trang_chu.html",
        {
            "request": request,
            "timestamp": int(datetime.now().timestamp())
        }
    )

@app.get("/game", response_class=HTMLResponse)
def game(request: Request):
    return templates.TemplateResponse(
        "game.html",
        {"request": request, "timestamp": request.state.timestamp}
    )

@app.get("/result", response_class=HTMLResponse)
def result(request: Request):
    return templates.TemplateResponse(
        "result.html",
        {"request": request, "timestamp": request.state.timestamp}
    )

@app.get("/random", response_class=HTMLResponse)
def random_mode(request: Request):
    return templates.TemplateResponse(
        "random.html",
        {"request": request, "timestamp": request.state.timestamp}
    )

@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request):
    return templates.TemplateResponse(
        "admin.html",
        {"request": request, "timestamp": request.state.timestamp}
    )

# -------------------- API QUIZZES --------------------
@app.get("/api/quizzes")
def get_quizzes():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM quizzes")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return JSONResponse(content=rows)

@app.post("/api/quizzes")
def create_quiz(data: dict):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO quizzes (name) VALUES (?)", (data["name"],))
    conn.commit()
    quiz_id = cur.lastrowid
    conn.close()
    return {"id": quiz_id}

@app.put("/api/quizzes/{quiz_id}")
def update_quiz(quiz_id: int, data: dict):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE quizzes SET name=? WHERE id=?", (data["name"], quiz_id))
    conn.commit()
    conn.close()
    return {"status": "updated"}

@app.delete("/api/quizzes/{quiz_id}")
def delete_quiz(quiz_id: int):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM quizzes WHERE id=?", (quiz_id,))
    conn.commit()
    conn.close()
    return {"status": "deleted"}

@app.get("/flashcard", response_class=HTMLResponse)
def flashcard_page(request: Request):
    return templates.TemplateResponse(
        "flashcard.html",
        {"request": request, "timestamp": request.state.timestamp}
    )

# -------------------- API QUESTIONS --------------------
@app.get("/api/questions/{quiz_id}")
def get_questions(quiz_id: int):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM questions WHERE quiz_id=?", (quiz_id,))
    rows = cur.fetchall()
    qs = []
    for r in rows:
        q = dict(r)
        if q["options"]:
            q["options"] = json.loads(q["options"])
        qs.append(q)
    conn.close()
    return JSONResponse(content=qs)

@app.get("/api/question/{q_id}")
def get_question(q_id: int):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM questions WHERE id=?", (q_id,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return {"error": "not found"}
    q = dict(row)
    q["question_image"] = q["question_image"]
    q["option_a_image"] = q["option_a_image"]
    q["option_b_image"] = q["option_b_image"]
    q["option_c_image"] = q["option_c_image"]
    q["option_d_image"] = q["option_d_image"]
    if q["options"]:
        q["options"] = json.loads(q["options"])
    return q

@app.post("/api/questions")
def create_question(data: dict):
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """INSERT INTO questions(quiz_id, question_text, question_type, question_image, options, correct_answer, option_a_image, option_b_image, option_c_image, option_d_image)
VALUES(?,?,?,?,?,?,?,?,?,?)""",
        ((data["quiz_id"], data["question_text"], data["question_type"], data.get("question_image"),
        json.dumps(data["options"]), data["correct_answer"], data.get("option_a_image"), data.get("option_b_image"), data.get("option_c_image"), data.get("option_d_image")))
    )
    conn.commit()
    qid = cur.lastrowid
    conn.close()
    return {"id": qid}

@app.put("/api/questions/{q_id}")
def update_question(q_id: int, data: dict):
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        """UPDATE questions SET question_text=?, question_type=?, question_image=?, options=?, correct_answer=?,
        option_a_image=?, option_b_image=?, option_c_image=?, option_d_image=? WHERE id=?""",
        ((data["question_text"], data["question_type"], data.get("question_image"), json.dumps(data["options"]), data["correct_answer"],
        data.get("option_a_image"), data.get("option_b_image"), data.get("option_c_image"), data.get("option_d_image"), q_id
        ))
    )
    conn.commit()
    conn.close()
    return {"status": "updated"}

@app.delete("/api/questions/{q_id}")
def delete_question(q_id: int):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM questions WHERE id=?", (q_id,))
    conn.commit()
    conn.close()
    return {"status": "deleted"}

GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")  # PAT lưu trong Render env
RENDER_TOKEN = os.getenv("RENDER_TOKEN")
REPO = "hoangphuc4913/VongQuayDiaLy"
BRANCH = "main"
RENDER_SERVICE_ID = "srv-d2hcps0gjchc73c2076g"

def commit_file_to_github(path, message):
    url = f"https://api.github.com/repos/{REPO}/contents/{path}"
    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",  # token, không phải Bearer
        "Accept": "application/vnd.github+json",
        "Content-Type": "application/json"
    }

    # lấy SHA cũ (nếu file đã tồn tại)
    r = requests.get(url, headers=headers)
    sha = None
    if r.status_code == 200:
        sha = r.json()["sha"]

    with open(path, "rb") as f:
        content = base64.b64encode(f.read()).decode()

    data = {
        "message": message,
        "content": content,
        "branch": BRANCH
    }
    if sha:
        data["sha"] = sha

    res = requests.put(url, headers=headers, json=data)
    print("GitHub status:", res.status_code)
    print("GitHub response:", res.text)
    return res.json()

@app.post("/api/save_quiz_file")
async def save_quiz_file():
    res_db = commit_file_to_github("quiz.db", "Update quiz.db")
    render_headers = {"Authorization": f"Bearer {RENDER_TOKEN}"}
    final_status = None

    async with httpx.AsyncClient() as client:
        while True:
            r = await client.get(
                f"https://api.render.com/v1/services/{RENDER_SERVICE_ID}/deploys",
                headers=render_headers,
            )
            data = r.json()
            deploy_status = data[0]["deploy"]["status"]
            print("Status: ", deploy_status)

            # Nếu deploy kết thúc (thành công hoặc fail hoặc bị hủy)
            if deploy_status in ["live", "canceled", "failed"]:
                final_status = deploy_status
                break

            # Nếu chưa xong thì chờ 5s check lại
            await asyncio.sleep(5)

    return {"status": final_status, "db":res_db}

#--Nhập xuất file excel--
@app.get("/api/export_questions/{quiz_id}")
def export_questions(quiz_id: int):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM questions WHERE quiz_id=?", (quiz_id,))
    rows = cur.fetchall()
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append([
        "question_text",
        "question_type",
        "A", "B", "C", "D",
        "correct_answer"
    ])
    for r in rows:
        q = dict(r)
        options = json.loads(q["options"]) if q["options"] else {}
        ws.append([
            q["question_text"],
            q["question_type"],
            options.get("A", ""),
            options.get("B", ""),
            options.get("C", ""),
            options.get("D", ""),
            q["correct_answer"]
        ])
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return Response(
        content=file_stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=quiz_{quiz_id}.xlsx"
        }
    )

@app.post("/api/import_questions/{quiz_id}")
async def import_questions(quiz_id: int, file: UploadFile = File(...)):
    content = await file.read()
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    conn = get_db()
    cur = conn.cursor()

    # Bỏ dòng header
    for row in ws.iter_rows(min_row=2, values_only=True):
        question_text, question_type, A, B, C, D, correct = row

        options = None
        if question_type != "note":
            options = {
                "A": A,
                "B": B,
                "C": C,
                "D": D
            }

        cur.execute("""
            INSERT INTO questions (quiz_id, question_text, question_type, options, correct_answer)
            VALUES (?, ?, ?, ?, ?)
        """, (
            quiz_id,
            question_text,
            question_type,
            json.dumps(options) if options else None,
            correct
        ))

    conn.commit()
    conn.close()

    return {"status": "imported"}

# -------------------- API IMAGES --------------------
@app.post("/api/upload_image")
async def upload_image(file: UploadFile = File(...)):

    ext = file.filename.split(".")[-1]

    filename = f"{uuid.uuid4()}.{ext}"

    save_path = f"static/uploads/{filename}"

    with open(save_path,"wb") as buffer:
        shutil.copyfileobj(file.file,buffer)

    return {
        "path":f"uploads/{filename}"
    }

